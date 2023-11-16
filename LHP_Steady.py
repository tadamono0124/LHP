# -*- coding: utf-8 -*-
"""
Created on Thu Jul  8 00:50:56 2021

@author: Satoshi Kajiyama
"""

# 大気環境下での定常解析

import numpy as np
import openpyxl
import LHP_Steady_module as LHP


###############################################################################
# 結果出力用ファイルの作成
###############################################################################
wb_filename = "LHP_result.xlsx"

wb = openpyxl.Workbook()
ws = wb["Sheet"]

ws.cell(row=1, column=1).value  = "Q_apply"
ws.cell(row=1, column=2).value  = "EC_energy"
ws.cell(row=1, column=3).value  = "CC_energy"
ws.cell(row=1, column=4).value  = "Q_hbec"
ws.cell(row=1, column=5).value  = "Q_ev"
ws.cell(row=1, column=6).value  = "Q_ecgr"
ws.cell(row=1, column=7).value  = "Q_wickccin"
ws.cell(row=1, column=8).value  = "Q_ecccc"
ws.cell(row=1, column=9).value  = "Q_sub"
ws.cell(row=1, column=10).value = "-Q_ccinccc"
ws.cell(row=1, column=11).value = "Q_hbamb"
ws.cell(row=1, column=12).value = "Q_ecamb"
ws.cell(row=1, column=13).value = "Q_cccamb"
ws.cell(row=1, column=14).value = "Q_out_vl"
ws.cell(row=1, column=15).value = "Q_out_cl"
ws.cell(row=1, column=16).value = "Q_out_ll"
ws.cell(row=1, column=17).value = "m_dot"
ws.cell(row=1, column=18).value = "dP_all"
ws.cell(row=1, column=19).value = "dP_capmax"
ws.cell(row=1, column=20).value = "V_cc"
ws.cell(row=1, column=21).value = "T_hb"
ws.cell(row=1, column=22).value = "T_ec"
ws.cell(row=1, column=23).value = "T_wickout"
ws.cell(row=1, column=24).value = "T_grout"
ws.cell(row=1, column=25).value = "T_ccc"
ws.cell(row=1, column=26).value = "T_ccin"
ws.cell(row=1, column=27).value = "T_vl_start"
ws.cell(row=1, column=28).value = "T_vl_end"
ws.cell(row=1, column=29).value = "T_vl_ave"
ws.cell(row=1, column=30).value = "T_cl_start"
ws.cell(row=1, column=31).value = "T_cl_end"
ws.cell(row=1, column=32).value = "T_cl_ave"
ws.cell(row=1, column=33).value = "T_ll_start"
ws.cell(row=1, column=34).value = "T_ll_end"
ws.cell(row=1, column=35).value = "T_ll_ave"
ws.cell(row=1, column=36).value = "P_wickout"
ws.cell(row=1, column=37).value = "P_grout"
ws.cell(row=1, column=38).value = "P_ccin"
ws.cell(row=1, column=39).value = "P_vl_start"
ws.cell(row=1, column=40).value = "P_vl_end"
ws.cell(row=1, column=41).value = "P_cl_start"
ws.cell(row=1, column=42).value = "P_cl_end"
ws.cell(row=1, column=43).value = "P_ll_start"
ws.cell(row=1, column=44).value = "P_ll_end"
ws.cell(row=1, column=45).value = "dP_wick"
ws.cell(row=1, column=46).value = "dP_gr"
ws.cell(row=1, column=47).value = "dP_vl"
ws.cell(row=1, column=48).value = "dP_cl"
ws.cell(row=1, column=49).value = "dP_ll"
ws.cell(row=1, column=50).value = "dP_all"
ws.cell(row=1, column=51).value = "h_evap"
ws.cell(row=1, column=52).value = "h_sink"
ws.cell(row=1, column=53).value = "T_sink"
ws.cell(row=1, column=54).value = "T_amb"
ws.cell(row=1, column=55).value = "N_evap"
ws.cell(row=1, column=56).value = "Q_ev_ratio"
ws.cell(row=1, column=57).value = "Re_gr"
ws.cell(row=1, column=58).value = "T_sur_ave"
ws.cell(row=1, column=59).value = "T_sur_ave_3"

ws.cell(row=1, column=60).value = "OD_vl"
ws.cell(row=1, column=61).value = "hi_cl"
ws.cell(row=1, column=62).value = "OD_ll"
ws.cell(row=1, column=63).value = "ID_vl"
ws.cell(row=1, column=64).value = "ID_cl"
ws.cell(row=1, column=65).value = "ID_ll"
ws.cell(row=1, column=66).value = "wick_l"
ws.cell(row=1, column=67).value = "wick_w"
ws.cell(row=1, column=68).value = "wick_h"
ws.cell(row=1, column=69).value = "HB_l"
ws.cell(row=1, column=70).value = "HB_w"
ws.cell(row=1, column=71).value = "EC_l"
ws.cell(row=1, column=72).value = "EC_w"
ws.cell(row=1, column=73).value = "EC_h"
ws.cell(row=1, column=74).value = "EC_t_x"
ws.cell(row=1, column=75).value = "EC_t_y"
ws.cell(row=1, column=76).value = "L_vl"
ws.cell(row=1, column=77).value = "L_cl"
ws.cell(row=1, column=78).value = "L_ll"
ws.cell(row=1, column=79).value = "N_vl"
ws.cell(row=1, column=80).value = "N_cl"
ws.cell(row=1, column=81).value = "N_ll"
ws.cell(row=1, column=82).value = "theta_vl"
ws.cell(row=1, column=83).value = "theta_con"
ws.cell(row=1, column=84).value = "theta_ll"
ws.cell(row=1, column=85).value = "cont_angle"
ws.cell(row=1, column=86).value = "r_pore"
ws.cell(row=1, column=87).value = "p_hi"
ws.cell(row=1, column=88).value = "K_wi"
ws.cell(row=1, column=89).value = "k_wick"
ws.cell(row=1, column=90).value = "k_cc"
ws.cell(row=1, column=91).value = "k_ec"
ws.cell(row=1, column=92).value = "k_hb"
ws.cell(row=1, column=93).value = "k_v"
ws.cell(row=1, column=94).value = "k_c"
ws.cell(row=1, column=95).value = "k_l"
ws.cell(row=1, column=96).value = "k_insu_hb"
ws.cell(row=1, column=97).value = "k_insu_ec"
ws.cell(row=1, column=98).value = "k_insu_cc"
ws.cell(row=1, column=99).value = "k_insu"
ws.cell(row=1, column=100).value = "h_out"
ws.cell(row=1, column=101).value = "h_cont_echb"
ws.cell(row=1, column=102).value = "h_cont_hb"
ws.cell(row=1, column=103).value = "h_cont_ec"
ws.cell(row=1, column=104).value = "h_cont_cc"
ws.cell(row=1, column=105).value = "h_cont"
ws.cell(row=1, column=106).value = "t_insu_hb"
ws.cell(row=1, column=107).value = "t_insu_ec"
ws.cell(row=1, column=108).value = "t_insu_cc"
ws.cell(row=1, column=109).value = "t_insu"


wb.save(wb_filename)


###############################################################################
# LHP諸元（インプットパラメータ）
###############################################################################

N_evap = 1                     # 蒸発器の個数 [-]


# フィッティングパラメータ
h_evap = 8000                # 蒸発熱伝達率 [W/m2/K]              ★
h_sink = 1000                  # 凝縮管における凝縮熱伝達率 [W/m2/K]　 ★

T_amb  = -35 + 273.15             # AmbientTemp. [K]　　　　　　　　　　      ★
T_sink = -35 + 273.15             # Sink Temp. [K]　　　　　　　　　　　      ★

OD_vl = 0.0025                 # OD of Vapor line [m]              ★
Hi_cl = 0.0020                 # Cold plateの厚さなので注意 [m]            ★
Wi_cl = 0.0040                 # Cold plateの各セクションの幅なので注意　[m]
OD_ll = 0.0025                 # OD. of liquid line [m]                   ★

ID_vl = 0.0012                # 蒸発管内径 [m]                    ★
ID_cl = 0.0012                # 凝縮管内径 [m]                    ★
ID_ll = 0.0012                # 液管内径 [m]                      ★

P_cri = 11333000.0             # 作動流体（アンモニア）の臨界圧 [Pa]　Shahの凝縮熱伝達のため

wick_l   = 0.048               # ウィックの横長さ[m]  ★
wick_w   = 0.0175               # ウィックの縦長さ[m]　 ★
wick_h   = 0.0007              # ウィックの厚さ[m]　　 ★

HB_l  = 0.050                  # ヒータブロックの横長さ[m]  ★
HB_w  = 0.019                  # ヒータブロックの縦長さ[m]  ★

EC_l   = 0.042                 # 蒸発器の外層の横長さ[m]  ★
EC_w   = 0.016                 # 蒸発器の外層の縦長さ[m]  ★
EC_h   = 0.0025                # 蒸発器の外層の厚さ[m]    ★
EC_t_x = 0.0010                # 蒸発器ケースの横方向厚さ[m]     ★
EC_t_y = 0.00075               # 蒸発器ケースの縦方向厚さ[m]     ★

CC_l      = 0.042                 # 補償器の外層の横長さ[m]  ★
CC_w      = 0.023                 # 補償器の外層の縦長さ[m]  ★ 
CC_h      = 0.0025                # 補償器の外層の厚さ[m]    ★
CC_t_x    = EC_t_x                # 補償器ケースの横方向厚さ[m]     ★
CC_t_y    = EC_t_y                # 補償器ケースの縦方向厚さ[m]     ★
CC_edge_h = 0.0006               # 補償器縁の高さ[m]     ★
dx_ecccc  = 0.01050               # 蒸発器縁と補償器中央との距離[m]     ★

l_g   = 0.0155                # 溝の軸方向長さ[m]     ★
t_g   = 0.0008                # 溝の高さ [m]          ★
w_g   = 0.0009                # 溝の幅 [m]           ★
#n_g   = 55
n_g   = wick_l / (2*w_g)      # グルーブとフィンの組合せ数 [-]  ★

L_vl  = 0.1622                 # 蒸発管長さ [m]  ★
L_cl  = 0.402                 # 凝縮管長さ [m]  ★
L_ll  = 0.100                 # 液管長さ [m]    ★

N_vl = 140                    # 蒸発管ノード分割数 [-]   ★
N_cl = 900                    # 凝縮管ノード分割数 [-]    ★
N_ll = 100                    # 液管ノード分割数 [-]     ★

theta_vl  = 0                 # 各種管の傾き [deg] ★
theta_con = 0                 # 各種管の傾き [deg] ★
theta_ll  = 0                 # 各種管の傾き [deg] ★

cont_angle  = 30                             # ウィック内の作動液の接触角 [deg]  ★ 
r_pore      = 9.6 * 1e-6                       # ウィックの細孔半径 [m]　　　　　★
p_hi        = 0.37                           # ウィックの空孔率 [-]　　　　　　　★
K_wi        = 2.7 *1e-13                     # ウィックの浸透率 [m^2]　　　　　　　★　
k_wick      = 7.5                           # ウィックの熱伝導率[W/m/K]          ★
                           
V_fluid    = 1.0*1e-3      # 作動流体の封入量 [L]

k_cc = 7.5                 # 補償器ケースの熱伝導率[W/m/K]    ★
k_ec = 7.5                 # 蒸発器ケースの熱伝導率[W/m/K]    ★
k_hb = 10                  # ヒータブロックの熱伝導率[W/m/K]　    ★　　　　　　
k_v  = 7.5                 # 蒸気管の熱伝導率 [W/m/K]        ★
k_c  = 7.5                 # 凝縮器の熱伝導率 [W/m/K]        ★
k_l  = 7.5                 # 液管の熱伝導率 [W/m/K]          ★

k_insu_hb = 0.0040           # 断熱材の熱伝導率（ヒータブロック）[W/m/K]　　
k_insu_ec = 0.0040           # 断熱材の熱伝導率（蒸発器ケース）[W/m/K]　　　
k_insu_cc = 0.0040           # 断熱材の熱伝導率（補償器ケース）[W/m/K]　　　
k_insu    = 0.0040    


       # 断熱材の熱伝導率（蒸気管と液管）[W/m/K]　　　


#############################   単位はすべて[W/m2/K]   ##################
h_out       = 10                  # 断熱材から外界への自然対流熱伝達          ★
h_cont_echb = 2000                # 接触熱伝達率（ヒータブロック→蒸発器ケース） ★

h_cont_hb   = 10          # 10000000 接触熱伝達率（ヒータブロック→断熱材）　     ★　　　
h_cont_ec   = 10           # 接触熱伝達率（蒸発器ケース→断熱材）　　     ★　　　
h_cont_cc   = 10            # 接触熱伝達率（補償器ケース→断熱材）　　     ★　　　
h_cont      = 10           # 接触熱伝達率（蒸気管＆液管→断熱材）         ★
                                                       
t_insu_hb = 0.010      # 断熱材の厚さ（ヒータブロック）[m]　         ★　　　　　
t_insu_ec = 0.010      # 断熱材の厚さ（蒸発器ケース）[m]　　         ★　　　　　
t_insu_cc = 0.005      # 断熱材の厚さ（補償器ケース）[m]　　         ★　　　　　
t_insu    = 0.015      # 蒸気管と液管の断熱材厚さ [m]               ★

###############################################################################
# 計算条件の設定
###############################################################################
Q_apply_str   = 0.5
dQ_apply      = 0.5 
Q_apply_end   = 10.0
T_wickout     = T_sink + 10          # (K)  初期値
  
iteration_CC = 10000000        # Twickoutのための反復計算回数
iteration_EC = 10000000        # mdのための反復計算回数

dT_wickout_stop = 1e-13 
dmd_stop        = 1e-13 


###############################################################################
###############################################################################

Q_apply_arange = np.arange(Q_apply_str, Q_apply_end+dQ_apply, dQ_apply)
n = 0                  # Q_applyカウンタ（結果出力用）


for Q_apply in Q_apply_arange:
    string = "Q_apply >>> " + str(Q_apply) + "W\n"
    print(string)

    p_CC_energy = 1

    dT_wickout = 1
    
    while np.abs(dT_wickout) > 1e-15:
        #print("dT_wickout : " + str(dT_wickout))
        
        p_EC_energy = -1
        md  = Q_apply / LHP.HV(T_wickout)
        dmd = -0.1 * md
        
        while np.abs(dmd) > 1e-15:
            #print("dmd : " + str(dmd))
            # 蒸発器における計算
            EVAP = LHP.Evap(Q_apply, md, T_wickout, h_evap, n_g, l_g, t_g, w_g, 
                            h_out, T_amb, h_cont_echb, h_cont_hb, t_insu_hb, k_insu_hb,
                            h_cont_ec, t_insu_ec, k_insu_ec, wick_l, HB_l, HB_w, 
                            EC_l, EC_w)
            
            P_wickout, P_grout, dP_gr = EVAP[0], EVAP[1], EVAP[2]
            T_ec, T_hb, T_grout = EVAP[3], EVAP[4], EVAP[5]
            Q_ev, Q_ecgr, Q_hbamb, Q_ecamb, Q_hbec = EVAP[6], EVAP[7], EVAP[8], EVAP[9], EVAP[10]
            Re_gr = EVAP[11]
            
            # 各種管における計算
            # 蒸気管に関する計算
            X_in = 1       # 蒸気相
            component = 1  # 1:蒸気管 2:凝縮器 3:液管
            VL = LHP.line_1D(md, P_grout, T_grout, T_amb, X_in, component, L_vl, N_vl, 
                             ID_vl, OD_vl, theta_vl, h_sink, k_v, k_c, k_l, 
                             t_insu, k_insu, h_cont, h_out, 0, P_cri)
            
            P_vl, T_vl, X_vl, V_vl = VL[0], VL[1], VL[2], VL[3]
            Q_out_vl = VL[4]
            Re_vl, Re_V_vl, Re_L_vl = VL[5], VL[6], VL[7]
            h_in_vl, reg_vl, Gcon_vl, dQ_out_vl = VL[8], VL[9], VL[10], VL[11]
            fai_v_r_vl, dPv2f_r_vl = VL[12], VL[13]
            hL_r_vl, hNu_r_vl, WeGT_r_vl, Z_r_vl, JG_r_vl = VL[14], VL[15], VL[16], VL[17], VL[18]   
            dPf_r_vl, dP2f_r_vl = VL[19], VL[20]
            dPh_r_vl, dPa_r_vl, dPall_r_vl, dPrat_r_vl = VL[21], VL[22], VL[23], VL[24]
            T_sur_r_vl = VL[25]
            
            
            # 凝縮管に関する計算
            component = 2  # 1:蒸気管 2:凝縮器 3:液管
            CL = LHP.line_1D(md, P_vl[-1], T_vl[-1], T_sink, X_vl[-1], component, 
                             L_cl, N_cl, ID_cl, Hi_cl, theta_con, h_sink, 
                             k_v, k_c, k_l, t_insu, k_insu, h_cont, h_out, Wi_cl, P_cri)
            
            #  コールドプレートに対応して，入力因数が変更されている
            #  OD_cl → hi_cl (液管と蒸気管はそのままOD_vlまたはOD_ll）
            #  wi_clも新たに追加（ただし，液管と蒸気管は通常の円管なのでwi_cl=0としている）
            #  Pcri 凝縮熱伝達式をShahにすることにより，作動流体の臨界圧が必要なため追加
            
            P_cl, T_cl, X_cl, V_cl = CL[0], CL[1], CL[2], CL[3]
            Q_out_cl = CL[4]
            Re_cl, Re_V_cl, Re_L_cl = CL[5], CL[6], CL[7]
            h_in_cl, reg_cl, Gcon_cl, dQ_out_cl = CL[8], CL[9], CL[10], CL[11]
            fai_v_r_cl, dPv2f_r_cl = CL[12], CL[13]
            hL_r_cl, hNu_r_cl, WeGT_r_cl, Z_r_cl, JG_r_cl = CL[14], CL[15], CL[16], CL[17], CL[18]   
            dPf_r_cl, dP2f_r_cl = CL[19], CL[20]
            dPh_r_cl, dPa_r_cl, dPall_r_cl, dPrat_r_cl = CL[21], CL[22], CL[23], CL[24]
            T_sur_r_cl = CL[25]
            
            
            # 液管に関する計算
            component = 3  # 1:蒸気管 2:凝縮器 3:液管
            LL = LHP.line_1D(md, P_cl[-1], T_cl[-1], T_amb, X_cl[-1], component,
                             L_ll, N_ll, ID_ll, OD_ll, theta_ll, h_sink, 
                             k_v, k_c, k_l, t_insu, k_insu, h_cont, h_out, 0, P_cri)
            
            P_ll, T_ll, X_ll, V_ll = LL[0], LL[1], LL[2], LL[3]
            Q_out_ll = LL[4]
            Re_ll, Re_V_ll, Re_L_ll = LL[5], LL[6], LL[7]
            h_in_ll, reg_ll, Gcon_ll, dQ_out_ll = LL[8], LL[9], LL[10], LL[11]
            fai_v_r_ll, dPv2f_r_ll = LL[12], LL[13]
            hL_r_ll, hNu_r_ll, WeGT_r_ll, Z_r_ll, JG_r_ll = LL[14], LL[15], LL[16], LL[17], LL[18]   
            dPf_r_ll, dP2f_r_ll = LL[19], LL[20]
            dPh_r_ll, dPa_r_ll, dPall_r_ll, dPrat_r_ll = LL[21], LL[22], LL[23], LL[24]
            T_sur_r_ll = LL[25]
            
            
            # CC内のボイド率:Vcc [-]の計算
            CCvoid = LHP.cc_void(P_ll, V_vl, V_cl, V_ll, T_wickout, T_vl, T_cl, T_ll,
                                 ID_vl, ID_cl, ID_ll, N_vl, N_cl, N_ll, L_vl, L_cl, L_ll,
                                 p_hi, EC_l, EC_w, EC_h, EC_t_x, EC_t_y, wick_l, wick_w, wick_h, 
                                 w_g, l_g, t_g, n_g, V_fluid, N_evap)
            
            V_cc, P_ccin, T_ccin = CCvoid[0], CCvoid[1], CCvoid[2]
            m_fluid, m_lines, m_evap, m_cc = CCvoid[3], CCvoid[4], CCvoid[5], CCvoid[6]
            
            Q_sub = LHP.Q_ccinll(md/N_evap, T_ll, T_ccin)
            
            ##　ここから要修正（4/25）
            # Gcccambの算出
            S_cc = 2 * CC_l * CC_w + 2 * CC_w * CC_h + CC_l * CC_h         # 蒸発器の側面と上面の表面積の和
            
            G_cccamb = 1 / ( 1 / (S_cc * h_cont_cc)
                        + t_insu_cc / (k_insu_cc * S_cc)
                        + 1 / (S_cc * h_out) )
                
            # Geccccの算出
            #D_out = 4 * (EC_l * EC_h) * (EC_w + EC_h) / (2 * (EC_l + EC_h) + 2 * (EC_w + EC_h))
            
            #D_in = 4 * (HB_l * HB_w) / (2 * (HB_l + HB_w))
            S_ecccc = CC_l * CC_h - (CC_l - 2 * CC_t_x) * (CC_h - 2 * CC_t_y)
            
            #G_ecccc = 2 * np.pi * k_ec * EC_t / np.log(D_out / D_in)  # 円管熱伝導として考えている．
            G_ecccc = k_ec * S_ecccc / dx_ecccc
            
            # Qeccccの算出
            S_ccin =  (CC_l - 2 * CC_t_x) * (CC_h - 2 * CC_t_y)                     # CC内を縦に切ったときの平均断面積
            L_ccin = 2 * (CC_l - 2 * CC_t_x) + 2 * (CC_h - 2 * CC_t_y)              # CC内を縦に切ったときの平均周囲長さ
            D_ccin = 4 * S_ccin / L_ccin                                            # CC内を縦に切ったときの平均断面の等価直径（mini_LHPノート参照）
            
            #　蒸発器内のボイド率は50#に固定
            h_ccinccc = 4.36 * (LHP.kL(T_ccin) * (1 - 0.5) + LHP.kV(T_ccin) * 0.5) / D_ccin     
            
            #補償器内の上面と側面の断面積を対象, ccc: 補償器ケース（金属）; ccin: 補償器内部（流体）
            G_ccinccc = (2 * ((CC_l - 2 * CC_t_x) + (CC_h - 2 * CC_t_y)) * (CC_w - CC_t_x)) * h_ccinccc
            
            T_ccc = ((G_ecccc * T_ec + G_cccamb * T_amb + G_ccinccc * T_ccin) 
                     / (G_ecccc + G_cccamb + G_ccinccc))
            
            Q_ecccc   = G_ecccc * (T_ec - T_ccc)
            
            # Qcccambの算出
            Q_cccamb  = G_cccamb * (T_ccc - T_amb)
            
            # Qccincccの算出
            Q_ccinccc = G_ccinccc * (T_ccin - T_ccc)
            
            # Qecwickleakの算出         
            #G_ecwickleak_sup =   # EC→wick間の熱抵抗を考慮
            #G_ecwickleak = 0.05 * LHP.lam_w(T_wickout, p_hi, k_wick) * wick_l * wick_w / wick_h
            
            G_ecwickleak = (1 / 
                            (1 / (LHP.lam_w(T_wickout, p_hi, k_wick) * wick_l * wick_w / wick_h)
                             + 1 / (k_ec * EC_t_y * EC_l / (dx_ecccc + 0.5 * wick_w))        
                             ))
            
            Q_ecwickleak = G_ecwickleak * (T_ec - T_ccin)
            Q_wickccin   = Q_ecwickleak
            
            
            #########　ここまで要修正
            
            # Qev_ratioの算出
            Q_ev_ratio = Q_ev / (Q_apply / N_evap)
            
            
            EC_energy = (Q_hbec - Q_ecamb - Q_ecccc - Q_ecgr - Q_ecwickleak - Q_ev) / Q_apply
            CC_energy = (Q_wickccin - Q_ccinccc - Q_sub) / Q_apply
            
            dP_capmax = 2 * LHP.surf_T(T_wickout) * np.cos(cont_angle * np.pi / 180) / r_pore
            
            dP_wick = ((md / N_evap) * LHP.MuL(T_wickout) * wick_h) / (LHP.LD(T_wickout) * K_wi * wick_l * wick_w)
            
            
            if EC_energy * p_EC_energy <= 0:
                dmd = -0.5 * dmd
            
            p_EC_energy = EC_energy
            
            if np.abs(dmd) < dmd_stop:
                break
            
            md = md + dmd
            
            if md < 0:
                break
        
        # dT_wickout
        if CC_energy * p_CC_energy <= 0:
            dT_wickout = -0.5 * dT_wickout
            
        p_CC_energy = CC_energy
        
        #print("Q_sub : " + str(Q_sub))
        #print("CC_energy : " + str(CC_energy))
        
        if np.abs(dT_wickout) < dT_wickout_stop:
            break
        
        T_wickout = T_wickout + dT_wickout
            
            

###############################################################################
    # 各熱負荷に対する結果の整理
###############################################################################
    
    # 圧力損失の整理
    # dPwick, dPcapmax, dPgr=Pwickout-Pgroutはすでに上で計算されているので省略．
    dP_vl = P_grout - P_vl[-1]
    dP_cl = P_vl[-1] - P_cl[-1]
    dP_ll = P_cl[-1] - P_ll[-1]
    
    dP_all   = P_wickout - P_ll[-1] + dP_wick
    dP_drive = dP_all / dP_capmax      # 1以下で駆動力あり
    
    
    # ファイル書き込み
    wb_LHP_line = openpyxl.Workbook()
    
    ws_vl = wb_LHP_line.worksheets[0]
    ws_vl.title = "Vapor Line"
    
    ws_vl.cell(row=1, column=1).value  = "P"
    ws_vl.cell(row=1, column=2).value  = "T"
    ws_vl.cell(row=1, column=3).value  = "X"
    ws_vl.cell(row=1, column=4).value  = "V"
    ws_vl.cell(row=1, column=5).value  = "Re"
    ws_vl.cell(row=1, column=6).value  = "Re(V)"
    ws_vl.cell(row=1, column=7).value  = "Re(L)"
    ws_vl.cell(row=1, column=8).value  = "h_in"
    ws_vl.cell(row=1, column=9).value  = "reg"
    ws_vl.cell(row=1, column=10).value = "Gcon"
    ws_vl.cell(row=1, column=11).value = "dQ_out"
    ws_vl.cell(row=1, column=12).value = "fai_v"
    ws_vl.cell(row=1, column=13).value = "dPv2f"
    ws_vl.cell(row=1, column=14).value = "hL"
    ws_vl.cell(row=1, column=15).value = "hNu"
    ws_vl.cell(row=1, column=16).value = "WeGT"
    ws_vl.cell(row=1, column=17).value = "Z"
    ws_vl.cell(row=1, column=18).value = "JG"
    ws_vl.cell(row=1, column=19).value = "dPf"
    ws_vl.cell(row=1, column=20).value = "dP2f"
    ws_vl.cell(row=1, column=21).value = "dPh"
    ws_vl.cell(row=1, column=22).value = "dPa"
    ws_vl.cell(row=1, column=23).value = "dPall"
    ws_vl.cell(row=1, column=24).value = "dPrat"
    ws_vl.cell(row=1, column=25).value = "T_sur"
    
    ws_cl = wb_LHP_line.copy_worksheet(wb_LHP_line["Vapor Line"])  
    ws_cl.title = "Condenser Line"
    
    ws_ll = wb_LHP_line.copy_worksheet(wb_LHP_line["Vapor Line"])  
    ws_ll.title = "Liquid Line"
    
    
    # Vapor Lineへのデータ書き込み
    for j in range(0, N_vl):
        ws_vl.cell(row=j+2, column=1).value  = P_vl[j]
        ws_vl.cell(row=j+2, column=2).value  = T_vl[j] - 273.15
        ws_vl.cell(row=j+2, column=3).value  = X_vl[j]
        ws_vl.cell(row=j+2, column=4).value  = V_vl[j]
        ws_vl.cell(row=j+2, column=5).value  = Re_vl[j] 
        ws_vl.cell(row=j+2, column=6).value  = Re_V_vl[j]
        ws_vl.cell(row=j+2, column=7).value  = Re_L_vl[j]
        ws_vl.cell(row=j+2, column=8).value  = h_in_vl[j] 
        ws_vl.cell(row=j+2, column=9).value  = reg_vl[j]
        ws_vl.cell(row=j+2, column=10).value = Gcon_vl[j]
        ws_vl.cell(row=j+2, column=11).value = dQ_out_vl[j]
        ws_vl.cell(row=j+2, column=12).value = fai_v_r_vl[j]
        ws_vl.cell(row=j+2, column=13).value = dPv2f_r_vl[j]
        ws_vl.cell(row=j+2, column=14).value = hL_r_vl[j]
        ws_vl.cell(row=j+2, column=15).value = hNu_r_vl[j]
        ws_vl.cell(row=j+2, column=16).value = WeGT_r_vl[j]
        ws_vl.cell(row=j+2, column=17).value = Z_r_vl[j]
        ws_vl.cell(row=j+2, column=18).value = JG_r_vl[j]
        ws_vl.cell(row=j+2, column=19).value = dPf_r_vl[j]
        ws_vl.cell(row=j+2, column=20).value = dP2f_r_vl[j]
        ws_vl.cell(row=j+2, column=21).value = dPh_r_vl[j]
        ws_vl.cell(row=j+2, column=22).value = dPa_r_vl[j]
        ws_vl.cell(row=j+2, column=23).value = dPall_r_vl[j]
        ws_vl.cell(row=j+2, column=24).value = dPrat_r_vl[j]
        ws_vl.cell(row=j+2, column=25).value = T_sur_r_vl[j] - 273.15
        
        
    # Condenser Lineへのデータ書き込み
    for j in range(0, N_cl):
        ws_cl.cell(row=j+2, column=1).value  = P_cl[j]
        ws_cl.cell(row=j+2, column=2).value  = T_cl[j] - 273.15
        ws_cl.cell(row=j+2, column=3).value  = X_cl[j]
        ws_cl.cell(row=j+2, column=4).value  = V_cl[j]
        ws_cl.cell(row=j+2, column=5).value  = Re_cl[j] 
        ws_cl.cell(row=j+2, column=6).value  = Re_V_cl[j]
        ws_cl.cell(row=j+2, column=7).value  = Re_L_cl[j]
        ws_cl.cell(row=j+2, column=8).value  = h_in_cl[j] 
        ws_cl.cell(row=j+2, column=9).value  = reg_cl[j]
        ws_cl.cell(row=j+2, column=10).value = Gcon_cl[j]
        ws_cl.cell(row=j+2, column=11).value = dQ_out_cl[j]
        ws_cl.cell(row=j+2, column=12).value = fai_v_r_cl[j]
        ws_cl.cell(row=j+2, column=13).value = dPv2f_r_cl[j]
        ws_cl.cell(row=j+2, column=14).value = hL_r_cl[j]
        ws_cl.cell(row=j+2, column=15).value = hNu_r_cl[j]
        ws_cl.cell(row=j+2, column=16).value = WeGT_r_cl[j]
        ws_cl.cell(row=j+2, column=17).value = Z_r_cl[j]
        ws_cl.cell(row=j+2, column=18).value = JG_r_cl[j]
        ws_cl.cell(row=j+2, column=19).value = dPf_r_cl[j]
        ws_cl.cell(row=j+2, column=20).value = dP2f_r_cl[j]
        ws_cl.cell(row=j+2, column=21).value = dPh_r_cl[j]
        ws_cl.cell(row=j+2, column=22).value = dPa_r_cl[j]
        ws_cl.cell(row=j+2, column=23).value = dPall_r_cl[j]
        ws_cl.cell(row=j+2, column=24).value = dPrat_r_cl[j]
        ws_cl.cell(row=j+2, column=25).value = T_sur_r_cl[j] - 273.15
    
    
    # Liquid Lineへのデータ書き込み
    for j in range(0, N_ll):
        ws_ll.cell(row=j+2, column=1).value  = P_ll[j]
        ws_ll.cell(row=j+2, column=2).value  = T_ll[j] - 273.15
        ws_ll.cell(row=j+2, column=3).value  = X_ll[j]
        ws_ll.cell(row=j+2, column=4).value  = V_ll[j]
        ws_ll.cell(row=j+2, column=5).value  = Re_ll[j] 
        ws_ll.cell(row=j+2, column=6).value  = Re_V_ll[j]
        ws_ll.cell(row=j+2, column=7).value  = Re_L_ll[j]
        ws_ll.cell(row=j+2, column=8).value  = h_in_ll[j] 
        ws_ll.cell(row=j+2, column=9).value  = reg_ll[j]
        ws_ll.cell(row=j+2, column=10).value = Gcon_ll[j]
        ws_ll.cell(row=j+2, column=11).value = dQ_out_ll[j]
        ws_ll.cell(row=j+2, column=12).value = fai_v_r_ll[j]
        ws_ll.cell(row=j+2, column=13).value = dPv2f_r_ll[j]
        ws_ll.cell(row=j+2, column=14).value = hL_r_ll[j]
        ws_ll.cell(row=j+2, column=15).value = hNu_r_ll[j]
        ws_ll.cell(row=j+2, column=16).value = WeGT_r_ll[j]
        ws_ll.cell(row=j+2, column=17).value = Z_r_ll[j]
        ws_ll.cell(row=j+2, column=18).value = JG_r_ll[j]
        ws_ll.cell(row=j+2, column=19).value = dPf_r_ll[j]
        ws_ll.cell(row=j+2, column=20).value = dP2f_r_ll[j]
        ws_ll.cell(row=j+2, column=21).value = dPh_r_ll[j]
        ws_ll.cell(row=j+2, column=22).value = dPa_r_ll[j]
        ws_ll.cell(row=j+2, column=23).value = dPall_r_ll[j]
        ws_ll.cell(row=j+2, column=24).value = dPrat_r_ll[j]
        ws_ll.cell(row=j+2, column=25).value = T_sur_r_ll[j] - 273.15
    
    
    # Excelブックの保存
    filepath_1 = "LHP_Line_"
    filepath_2 = str(Q_apply)
    filepath_3 = "W.xlsx"
    filepath_line = filepath_1 + filepath_2 + filepath_3
    
    wb_LHP_line.save(filepath_line)


    T_vl_ave  = np.mean(T_vl)
    T_cl_ave  = np.mean(T_cl)
    T_ll_ave  = np.mean(T_ll)
    T_sur_ave = np.mean(T_sur_r_cl)

    T_sur_ave_3 = (T_sur_r_cl[0] + T_sur_r_cl[int(N_cl/2)] + T_sur_r_cl[-1]) / 3
    
    
    wb_LHP = openpyxl.load_workbook(wb_filename)
    ws_LHP = wb_LHP["Sheet"]
    
    ws_LHP.cell(row=n+2, column=1).value  = Q_apply
    ws_LHP.cell(row=n+2, column=2).value  = EC_energy
    ws_LHP.cell(row=n+2, column=3).value  = CC_energy
    ws_LHP.cell(row=n+2, column=4).value  = Q_hbec
    ws_LHP.cell(row=n+2, column=5).value  = Q_ev
    ws_LHP.cell(row=n+2, column=6).value  = Q_ecgr
    ws_LHP.cell(row=n+2, column=7).value  = Q_wickccin
    ws_LHP.cell(row=n+2, column=8).value  = Q_ecccc
    ws_LHP.cell(row=n+2, column=9).value  = Q_sub
    ws_LHP.cell(row=n+2, column=10).value = -Q_ccinccc
    ws_LHP.cell(row=n+2, column=11).value = Q_hbamb
    ws_LHP.cell(row=n+2, column=12).value = Q_ecamb
    ws_LHP.cell(row=n+2, column=13).value = Q_cccamb
    ws_LHP.cell(row=n+2, column=14).value = Q_out_vl
    ws_LHP.cell(row=n+2, column=15).value = Q_out_cl
    ws_LHP.cell(row=n+2, column=16).value = Q_out_ll
    ws_LHP.cell(row=n+2, column=17).value = md
    ws_LHP.cell(row=n+2, column=18).value = dP_all
    ws_LHP.cell(row=n+2, column=19).value = dP_capmax
    ws_LHP.cell(row=n+2, column=20).value = V_cc
    ws_LHP.cell(row=n+2, column=21).value = T_hb - 273.15
    ws_LHP.cell(row=n+2, column=22).value = T_ec - 273.15
    ws_LHP.cell(row=n+2, column=23).value = T_wickout - 273.15
    ws_LHP.cell(row=n+2, column=24).value = T_grout - 273.15
    ws_LHP.cell(row=n+2, column=25).value = T_ccc - 273.15
    ws_LHP.cell(row=n+2, column=26).value = T_ccin - 273.15
    ws_LHP.cell(row=n+2, column=27).value = T_vl[0] - 273.15
    ws_LHP.cell(row=n+2, column=28).value = T_vl[-1] - 273.15
    ws_LHP.cell(row=n+2, column=29).value = T_vl_ave - 273.15
    ws_LHP.cell(row=n+2, column=30).value = T_cl[0] - 273.15
    ws_LHP.cell(row=n+2, column=31).value = T_cl[-1] - 273.15
    ws_LHP.cell(row=n+2, column=32).value = T_cl_ave - 273.15
    ws_LHP.cell(row=n+2, column=33).value = T_ll[0] - 273.15
    ws_LHP.cell(row=n+2, column=34).value = T_ll[-1] - 273.15
    ws_LHP.cell(row=n+2, column=35).value = T_ll_ave - 273.15
    ws_LHP.cell(row=n+2, column=36).value = P_wickout
    ws_LHP.cell(row=n+2, column=37).value = P_grout
    ws_LHP.cell(row=n+2, column=38).value = P_ccin
    ws_LHP.cell(row=n+2, column=39).value = P_vl[0]
    ws_LHP.cell(row=n+2, column=40).value = P_vl[-1]
    ws_LHP.cell(row=n+2, column=41).value = P_cl[0]
    ws_LHP.cell(row=n+2, column=42).value = P_cl[-1]
    ws_LHP.cell(row=n+2, column=43).value = P_ll[0]
    ws_LHP.cell(row=n+2, column=44).value = P_ll[-1]
    ws_LHP.cell(row=n+2, column=45).value = dP_wick
    ws_LHP.cell(row=n+2, column=46).value = dP_gr
    ws_LHP.cell(row=n+2, column=47).value = dP_vl
    ws_LHP.cell(row=n+2, column=48).value = dP_cl
    ws_LHP.cell(row=n+2, column=49).value = dP_ll
    ws_LHP.cell(row=n+2, column=50).value = h_evap
    ws_LHP.cell(row=n+2, column=51).value = h_sink
    ws_LHP.cell(row=n+2, column=52).value = T_sink - 273.15
    ws_LHP.cell(row=n+2, column=53).value = T_amb - 273.15
    ws_LHP.cell(row=n+2, column=54).value = N_evap
    ws_LHP.cell(row=n+2, column=55).value = Q_ev_ratio
    ws_LHP.cell(row=n+2, column=56).value = Re_gr
    ws_LHP.cell(row=n+2, column=57).value = T_sur_ave - 273.15
    ws_LHP.cell(row=n+2, column=58).value = T_sur_ave_3 - 273.15
    
    ws_LHP.cell(row=n+2, column=60).value = OD_vl
    ws_LHP.cell(row=n+2, column=61).value = Hi_cl
    ws_LHP.cell(row=n+2, column=62).value = OD_ll
    ws_LHP.cell(row=n+2, column=63).value = ID_vl
    ws_LHP.cell(row=n+2, column=64).value = ID_cl
    ws_LHP.cell(row=n+2, column=65).value = ID_ll
    ws_LHP.cell(row=n+2, column=66).value = wick_l
    ws_LHP.cell(row=n+2, column=67).value = wick_w
    ws_LHP.cell(row=n+2, column=68).value = wick_h
    ws_LHP.cell(row=n+2, column=69).value = HB_l
    ws_LHP.cell(row=n+2, column=70).value = HB_w
    ws_LHP.cell(row=n+2, column=71).value = EC_l
    ws_LHP.cell(row=n+2, column=72).value = EC_w
    ws_LHP.cell(row=n+2, column=73).value = EC_h
    ws_LHP.cell(row=n+2, column=74).value = EC_t_x
    ws_LHP.cell(row=n+2, column=75).value = EC_t_y
    ws_LHP.cell(row=n+2, column=76).value = L_vl
    ws_LHP.cell(row=n+2, column=77).value = L_cl
    ws_LHP.cell(row=n+2, column=78).value = L_ll
    ws_LHP.cell(row=n+2, column=79).value = N_vl
    ws_LHP.cell(row=n+2, column=80).value = N_cl
    ws_LHP.cell(row=n+2, column=81).value = N_ll
    ws_LHP.cell(row=n+2, column=82).value = theta_vl
    ws_LHP.cell(row=n+2, column=83).value = theta_con
    ws_LHP.cell(row=n+2, column=84).value = theta_ll
    ws_LHP.cell(row=n+2, column=85).value = cont_angle
    ws_LHP.cell(row=n+2, column=86).value = r_pore
    ws_LHP.cell(row=n+2, column=87).value = p_hi
    ws_LHP.cell(row=n+2, column=88).value = K_wi
    ws_LHP.cell(row=n+2, column=89).value = k_wick
    ws_LHP.cell(row=n+2, column=90).value = k_cc
    ws_LHP.cell(row=n+2, column=91).value = k_ec
    ws_LHP.cell(row=n+2, column=92).value = k_hb
    ws_LHP.cell(row=n+2, column=93).value = k_v
    ws_LHP.cell(row=n+2, column=94).value = k_c
    ws_LHP.cell(row=n+2, column=95).value = k_l
    ws_LHP.cell(row=n+2, column=96).value = k_insu_hb
    ws_LHP.cell(row=n+2, column=97).value = k_insu_ec
    ws_LHP.cell(row=n+2, column=98).value = k_insu_cc
    ws_LHP.cell(row=n+2, column=99).value = k_insu
    ws_LHP.cell(row=n+2, column=100).value = h_out
    ws_LHP.cell(row=n+2, column=101).value = h_cont_echb
    ws_LHP.cell(row=n+2, column=102).value = h_cont_hb
    ws_LHP.cell(row=n+2, column=103).value = h_cont_ec
    ws_LHP.cell(row=n+2, column=104).value = h_cont_cc
    ws_LHP.cell(row=n+2, column=105).value = h_cont
    ws_LHP.cell(row=n+2, column=106).value = t_insu_hb
    ws_LHP.cell(row=n+2, column=107).value = t_insu_ec
    ws_LHP.cell(row=n+2, column=108).value = t_insu_cc
    ws_LHP.cell(row=n+2, column=109).value = t_insu
    
    wb_LHP.save(wb_filename)
    
    n = n + 1
        
###############################################################################
    #　各熱負荷に対するキャピラリー限界の設定
###############################################################################

    if dP_all > dP_capmax:
        print("LHP is dry out")
        break


###############################################################################
    #　グラフ描画
###############################################################################
LHP.graphing(wb_filename)


















